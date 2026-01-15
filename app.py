
# file: app.py
from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path
from typing import List, Dict

import pandas as pd
import streamlit as st

from tiss_parser import (
    parse_tiss_xml,
    parse_many_xmls,
    audit_por_guia,
    __version__ as PARSER_VERSION
)

st.set_page_config(page_title="Leitor TISS XML", layout="wide")
st.title("Leitor de XML TISS (Consulta e SP‑SADT)")
st.caption(f"Extrai nº do lote, quantidade de guias e valor total por arquivo • Parser {PARSER_VERSION}")

tab1, tab2 = st.tabs(["Upload de XML(s)", "Ler de uma pasta local (clonada do GitHub)"])

# =========================================================
# FORMATAÇÃO DE MOEDA (BR)
# =========================================================
def format_currency_br(val) -> str:
    """Converte número em string 'R$ 1.234,56'. Valores inválidos viram 'R$ 0,00'."""
    try:
        v = float(Decimal(str(val)))
    except Exception:
        v = 0.0
    neg = v < 0
    v = abs(v)
    inteiro = int(v)
    centavos = int(round((v - inteiro) * 100))
    inteiro_fmt = f"{inteiro:,}".replace(",", ".")
    centavos_fmt = f"{centavos:02d}"
    s = f"R$ {inteiro_fmt},{centavos_fmt}"
    return f"-{s}" if neg else s


def _df_display_currency(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """
    Retorna uma cópia do df com as colunas em 'cols' formatadas como moeda BR (texto),
    apenas para exibição na tela (mantemos o df original numérico para cálculos/exports).
    """
    dfd = df.copy()
    for c in cols:
        if c in dfd.columns:
            dfd[c] = dfd[c].apply(format_currency_br)
    return dfd


# =========================================================
# Utils de dataframe/saída
# =========================================================
def _to_float(val) -> float:
    try:
        return float(Decimal(str(val)))
    except Exception:
        return 0.0

def _df_format(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garantia de tipos e ordenação para exibição/CSV (numérico).
    Acrescenta coluna 'suspeito' quando qtde_guias>0 e valor_total==0.
    """
    if 'valor_total' in df.columns:
        df['valor_total'] = df['valor_total'].apply(_to_float)
    if 'qtde_guias' in df.columns and 'valor_total' in df.columns:
        df['suspeito'] = (df['qtde_guias'] > 0) & (df['valor_total'] == 0)
    else:
        df['suspeito'] = False

    if 'erro' not in df.columns:
        df['erro'] = None

    ordenar = ['numero_lote', 'tipo', 'qtde_guias', 'valor_total', 'estrategia_total',
               'arquivo', 'suspeito', 'erro', 'parser_version']
    cols = [c for c in ordenar if c in df.columns] + [c for c in df.columns if c not in ordenar]
    df = df[cols].sort_values(['numero_lote', 'tipo', 'arquivo'], ignore_index=True)
    return df

def _make_agg(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=['numero_lote', 'tipo', 'qtde_arquivos', 'qtde_guias_total', 'valor_total'])
    agg = df.groupby(['numero_lote', 'tipo'], dropna=False, as_index=False).agg(
        qtde_arquivos=('arquivo', 'count'),
        qtde_guias_total=('qtde_guias', 'sum'),
        valor_total=('valor_total', 'sum')
    ).sort_values(['numero_lote', 'tipo'], ignore_index=True)
    return agg

def _download_excel_button(df_resumo: pd.DataFrame, df_agg: pd.DataFrame, df_auditoria: pd.DataFrame, label: str):
    """
    Gera um Excel com formatação 'R$ #,##0.00' nas colunas monetárias.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        (df_resumo if not df_resumo.empty else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="Resumo por arquivo"
        )
        (df_agg if not df_agg.empty else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="Agregado por lote"
        )
        (df_auditoria if not df_auditoria.empty else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="Auditoria"
        )

        # Aplicar formato moeda via openpyxl
        from openpyxl.styles import numbers

        def format_currency_sheet(ws, header_row=1, currency_cols=("valor_total", "total_tag",
                                                                   "subtotal_itens_proc", "subtotal_itens_outras",
                                                                   "subtotal_itens")):
            headers = {ws.cell(row=header_row, column=c).value: c for c in range(1, ws.max_column + 1)}
            numfmt = 'R$ #,##0.00'
            for col_name in currency_cols:
                if col_name in headers:
                    col_idx = headers[col_name]
                    for r in range(header_row + 1, ws.max_row + 1):
                        ws.cell(row=r, column=col_idx).number_format = numfmt

        # Resumo
        ws = writer.sheets.get("Resumo por arquivo")
        if ws is not None:
            format_currency_sheet(ws, currency_cols=("valor_total",))
        # Agregado
        ws = writer.sheets.get("Agregado por lote")
        if ws is not None:
            format_currency_sheet(ws, currency_cols=("valor_total",))
        # Auditoria
        ws = writer.sheets.get("Auditoria")
        if ws is not None:
            format_currency_sheet(ws, currency_cols=("total_tag","subtotal_itens_proc","subtotal_itens_outras","subtotal_itens"))

    st.download_button(
        label,
        data=buffer.getvalue(),
        file_name="resumo_xml_tiss.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def _auditar_alertas(df: pd.DataFrame) -> None:
    if df.empty:
        return
    sus = df[df['suspeito']]
    err = df[df['erro'].notna()] if 'erro' in df.columns else pd.DataFrame()

    if not sus.empty:
        st.warning(
            f"⚠️ {len(sus)} arquivo(s) com valor_total=0 e qtde_guias>0. Verifique: "
            + ", ".join(sus['arquivo'].tolist())[:500]
        )
    if not err.empty:
        st.error(
            f"❌ {len(err)} arquivo(s) com erro no parsing. Exemplos: "
            + ", ".join(err['arquivo'].head(5).tolist())
        )

# =========================================================
# Upload
# =========================================================
with tab1:
    files = st.file_uploader(
        "Selecione um ou mais arquivos XML TISS",
        type=['xml'],
        accept_multiple_files=True
    )
    if files:
        resultados: List[Dict] = []
        for f in files:
            try:
                if hasattr(f, "seek"):
                    f.seek(0)
                res = parse_tiss_xml(f)
                res['arquivo'] = f.name
                if 'erro' not in res:
                    res['erro'] = None
            except Exception as e:
                res = {
                    'arquivo': f.name,
                    'numero_lote': '',
                    'tipo': 'DESCONHECIDO',
                    'qtde_guias': 0,
                    'valor_total': Decimal('0'),
                    'estrategia_total': 'erro',
                    'parser_version': PARSER_VERSION,
                    'erro': str(e),
                }
            resultados.append(res)

        if resultados:
            df = pd.DataFrame(resultados)
            df = _df_format(df)  # numérico para cálculos
            df_disp = _df_display_currency(df, ['valor_total'])  # cópia apenas para exibir em R$

            st.subheader("Resumo por arquivo")
            st.dataframe(df_disp, use_container_width=True)

            st.subheader("Agregado por nº do lote")
            agg = _make_agg(df)
            agg_disp = _df_display_currency(agg, ['valor_total'])
            st.dataframe(agg_disp, use_container_width=True)

            _auditar_alertas(df)

            col1, col2, col3 = st.columns(3)
            with col1:
                # CSV numérico (recomendado)
                st.download_button(
                    "Baixar resumo (CSV)",
                    df.to_csv(index=False).encode('utf-8'),
                    file_name="resumo_xml_tiss.csv",
                    mime="text/csv"
                )
            with col2:
                # Excel com formatação BR nas abas
                _download_excel_button(df, agg, df, "Baixar resumo (Excel .xlsx)")
            with col3:
                st.caption("O Excel inclui as abas: Resumo, Agregado e Auditoria (moeda BR).")

            with st.expander("🔎 Auditoria por guia (opcional)"):
                arquivo_escolhido = st.selectbox("Selecione um arquivo enviado", options=[r['arquivo'] for r in resultados])
                if st.button("Gerar auditoria do arquivo selecionado"):
                    escolhido = next((f for f in files if f.name == arquivo_escolhido), None)
                    if escolhido is not None:
                        if hasattr(escolhido, "seek"):
                            escolhido.seek(0)
                        linhas = audit_por_guia(escolhido)
                        df_a = pd.DataFrame(linhas)
                        # Mantém numérico internamente; cria uma cópia para exibir em BR
                        df_a_disp = df_a.copy()
                        for c in ('total_tag', 'subtotal_itens_proc', 'subtotal_itens_outras', 'subtotal_itens'):
                            if c in df_a_disp.columns:
                                df_a_disp[c] = df_a_disp[c].apply(format_currency_br)
                        st.dataframe(df_a_disp, use_container_width=True)
                        # CSV de auditoria: numérico (melhor para conferência/BI)
                        st.download_button(
                            "Baixar auditoria (CSV)",
                            df_a.to_csv(index=False).encode('utf-8'),
                            file_name=f"auditoria_{arquivo_escolhido}.csv",
                            mime="text/csv"
                        )

# =========================================================
# Pasta local (útil para rodar local/clonado)
# =========================================================
with tab2:
    pasta = st.text_input(
        "Caminho da pasta com XMLs (ex.: ./data ou C:\\repos\\tiss-xmls)",
        value="./data"
    )
    if st.button("Ler pasta"):
        p = Path(pasta)
        if not p.exists():
            st.error("Pasta não encontrada.")
        else:
            xmls = list(p.glob("*.xml"))
            if not xmls:
                st.warning("Nenhum .xml encontrado nessa pasta.")
            else:
                resultados = parse_many_xmls(xmls)
                df = pd.DataFrame(resultados)
                df = _df_format(df)
                df_disp = _df_display_currency(df, ['valor_total'])

                st.subheader("Resumo por arquivo")
                st.dataframe(df_disp, use_container_width=True)

                st.subheader("Agregado por nº do lote")
                agg = _make_agg(df)
                agg_disp = _df_display_currency(agg, ['valor_total'])
                st.dataframe(agg_disp, use_container_width=True)

                _auditar_alertas(df)

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.download_button(
                        "Baixar resumo (CSV)",
                        df.to_csv(index=False).encode('utf-8'),
                        file_name="resumo_xml_tiss.csv",
                        mime="text/csv"
                    )
                with col2:
                    _download_excel_button(df, agg, df, "Baixar resumo (Excel .xlsx)")
                with col3:
                    st.caption("O Excel inclui as abas: Resumo, Agregado e Auditoria (moeda BR).")
